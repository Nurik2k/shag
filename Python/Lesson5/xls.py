import openpyxl
import clipboard as cb
import os.path


def wb_init() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    return wb


def wb_open(filename: str) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(filename)
    return wb


def wb_save(filename: str, wb: openpyxl.Workbook) -> None:
    wb.save(filename)


# def wb_sheet_create(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet:
#     return wb.create_sheet(name)


def wb_get_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet:
    return wb[name]


def main():
    # Проверка на существование файла
    if not os.path.exists("test.xlsx"):
        wb = wb_init()
        wb_save("test.xlsx", wb)

    wb = wb_open("test.xlsx")

    print(f"{wb.sheetnames[0]}")

    sheet = wb_get_sheet(wb, wb.sheetnames[0])

    clipbard_sentence = cb.get_from_clipboard()

    for row_number in range(1, sheet.max_row + 1):
        print(sheet.cell(row=row_number, column=1).value)
        sheet.cell(row=row_number, column=1).value = clipbard_sentence

    # Осуществить проверку на существование строки из буфера обмена в ячейке
    # Если такой строки нет, то добавить ее.

    wb_save("test.xlsx", wb)


if __name__ == "__main__":
    main()

    # * Содать буфер обмена из 20 строк из буфера обмена
    # Выводить пользователю этот буфер обмена с нумерацией
    # Попросить польщзователя ввести номер, который он хочет поместить в буфер обмена
    # Поместить в буфер обмена строку с этим номером
